#!/usr/bin/env python3
"""
Parses the OnCall Excel schedule and writes the CURRENT week's rotation
(all three shifts, holidays, swap notes, and each person's contact info)
to a small JSON file the website reads.

Usage:
    python3 parse_oncall.py /path/to/WebHosting_OnCall.xlsx /path/to/output/oncall.json

Intended to run on a schedule (cron) — see the bottom of this file for a
sample crontab line. Each run overwrites the JSON output with the current
week's data; nothing here modifies the source spreadsheet.

Expected spreadsheet layout ("On-Call Schedule" sheet):
    Row 1 (merged headers): Date | India (7:00 PM - 2:00 AM CST) | ... |
                             India (2:00 AM - 9:00 AM CST) | ... |
                             USA (9:00 AM - 7:00 PM CST) | ... |
                             Holidays | Swap Notes   (columns H/I — internal
                             use only, intentionally not read by this script)
    Row 2 (sub-headers):    (blank) | Primary | Secondary | Primary |
                             Secondary | Primary | Secondary | (blank) | (blank)
    Row 3+:                 one row per week, "Date" = the Monday that
                             week starts

Expected "Contacts" sheet: Name | Mobile | Location
"""

import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "Missing dependency: openpyxl.\n"
        "Install with: pip3 install openpyxl --break-system-packages\n"
        "(or: pip3 install --user openpyxl, depending on your environment)\n"
    )
    sys.exit(1)

SCHEDULE_SHEET = "On-Call Schedule"
CONTACTS_SHEET = "Contacts"

SHIFT_COLUMNS = [
    # (label, primary_col_index, secondary_col_index) — 0-indexed into each row tuple
    ("India (7:00 PM - 2:00 AM CST)", 1, 2),
    ("India (2:00 AM - 9:00 AM CST)", 3, 4),
    ("USA (9:00 AM - 7:00 PM CST)", 5, 6),
]


def load_contacts(wb):
    """Returns {name: {"mobile": ..., "location": ...}}"""
    contacts = {}
    if CONTACTS_SHEET not in wb.sheetnames:
        return contacts
    ws = wb[CONTACTS_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row
    for row in rows:
        if not row or not row[0]:
            continue
        name, mobile, location = (list(row) + [None, None, None])[:3]
        contacts[str(name).strip()] = {
            "mobile": mobile,
            "location": location,
        }
    return contacts


def find_current_week_row(ws, today):
    """
    Each data row's Date is the Monday that week starts. The "current"
    row is the one whose week [Date, Date+7) contains today.
    """
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # skip both header rows
    candidates = []
    for row in rows:
        week_start = row[0]
        if not isinstance(week_start, datetime):
            continue
        week_start = week_start.date() if hasattr(week_start, "date") else week_start
        candidates.append((week_start, row))

    candidates.sort(key=lambda c: c[0])

    for week_start, row in candidates:
        week_end = week_start + timedelta(days=7)
        if week_start <= today < week_end:
            return week_start, row

    return None, None


def person(name, contacts):
    if not name:
        return None
    name = str(name).strip()
    info = contacts.get(name, {})
    return {
        "name": name,
        "mobile": info.get("mobile"),
        "location": info.get("location"),
    }


def main():
    if len(sys.argv) != 3:
        sys.stderr.write(f"Usage: {sys.argv[0]} <input.xlsx> <output.json>\n")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not input_path.exists():
        sys.stderr.write(f"Input file not found: {input_path}\n")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, data_only=True)

    if SCHEDULE_SHEET not in wb.sheetnames:
        sys.stderr.write(
            f"Sheet '{SCHEDULE_SHEET}' not found. Sheets present: {wb.sheetnames}\n"
        )
        sys.exit(1)

    ws = wb[SCHEDULE_SHEET]
    contacts = load_contacts(wb)

    today = datetime.now().date()
    week_start, row = find_current_week_row(ws, today)

    if row is None:
        sys.stderr.write(
            f"No schedule row found covering today ({today}). "
            "The spreadsheet may not yet include this week, or the date "
            "range has passed — check the source file.\n"
        )
        # Still write a JSON file so the site can show a clear "no data"
        # state instead of silently keeping stale data forever.
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps({
            "weekOf": None,
            "shifts": [],
            "updatedAt": datetime.now(timezone.utc).isoformat(),
            "error": "No schedule row found for the current week",
        }, indent=2))
        sys.exit(1)

    shifts = []
    for label, p_idx, s_idx in SHIFT_COLUMNS:
        shifts.append({
            "label": label,
            "primary": person(row[p_idx], contacts),
            "secondary": person(row[s_idx], contacts),
        })

    output = {
        "weekOf": week_start.isoformat(),
        "shifts": shifts,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2))
    print(f"Wrote {output_path} for week of {week_start.isoformat()}")


if __name__ == "__main__":
    main()

# ---------------------------------------------------------------------------
# Sample crontab entry — refresh every weekday morning at 6:00 AM:
#
#   0 6 * * 1-5 /usr/bin/python3 /path/to/parse_oncall.py \
#       /path/to/synced/WebHosting_OnCall_2026.xlsx \
#       /ebiz/tomcat/webapps/ROOT/oncall.json \
#       >> /var/log/oncall-refresh.log 2>&1
#
# IMPORTANT CAVEAT: /ebiz/tomcat/webapps/ROOT is the LOCALLY UNPACKED copy
# of ROOT.war. If Tomcat redeploys/re-unpacks the WAR (e.g. after copying
# a new version to the NAS docBase), this file will be wiped until the
# next cron run. Either re-run this script manually right after any
# redeploy, or — better long-term — mount an external, persistent
# directory into the Context via a <Resources><PostResources .../></Resources>
# block in ROOT.xml, so this JSON lives outside the WAR's exploded
# directory entirely and survives redeploys automatically. Worth setting
# up if this becomes a recurring annoyance.
# ---------------------------------------------------------------------------
